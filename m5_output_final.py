import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

ordem = {"SEM_MATCH": 0, "REVISAR": 1, "MATCH": 2}
cores = {"MATCH": "C6EFCE", "REVISAR": "FFEB9C", "SEM_MATCH": "FFC7CE"}

df = pd.read_excel("resultado_matching.xlsx")
df["ordem"] = df["status"].map(ordem)
df = df.sort_values("ordem").drop(columns="ordem").reset_index(drop=True)
df.to_excel("resultado_final.xlsx", index=False)

wb = load_workbook("resultado_final.xlsx")
ws = wb.active

col_status = [c.column for c in ws[1] if c.value == "status"][0]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    status = row[col_status - 1].value
    fill = PatternFill("solid", fgColor=cores.get(status, "FFFFFF"))
    for cell in row:
        cell.fill = fill

wb.save("resultado_final.xlsx")
print("Output gerado → resultado_final.xlsx")
